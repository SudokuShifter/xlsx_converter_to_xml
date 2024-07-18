from openpyxl import load_workbook
from yattag import Doc, indent

wb = load_workbook('table1.xlsx')

ws = wb.worksheets[0]

doc, tag, text = Doc().tagtext()

xml_schema = '<СчетаПК xmlns="http://v8.1c.ru/edi/edi_stnd/109" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" ДатаФормирования="2024-03-22" НомерДоговора="1434171" НаименованиеОрганизации="ООО Диал-Авто" ИНН="2128051718" ИдПервичногоДокумента="00000000-0000-0000-0000-000000000000" xsi:schemaLocation="http://v8.1c.ru/edi/edi_stnd/109 Wages_3.2.xsd">'
doc.asis(xml_schema)

# Итерация по строкам Excel-файла
for row in ws.iter_rows(min_row=1, max_col=5, values_only=True):
    with tag('РезультатОткрытияСчетов'):
        with tag('Сотрудник', 'Нпп="1"'):
            with tag('Фамилия'):
                text(row[0] if row[0] else '')
            with tag('Имя'):
                text(row[1] if row[1] else '')
            with tag('Отчество'):
                text(row[2] if row[2] else '')
            with tag('ЛицевойСчет'):
                text(''.join(row[4].split()) if row[4] else '')
            with tag('УдостоверениеЛичности'):
                with tag('КодВидаДокумента'):
                    text('')
                with tag('ВидДокумента'):
                    text('')
                with tag('Серия'):
                    text('')
                with tag('Номер'):
                    text('')
                with tag('ДатаВыдачи'):
                    text('')
                with tag('КемВыдан'):
                    text('')

    with tag('КонтрольныеСуммы'):
        with tag('КоличествоЗаписей'):
            text('')
        with tag('СуммаИтого'):
            text('')

# Завершение корневого элемента
doc.asis('</СчетаПК>')

# Форматирование и запись результата в файл
result = indent(doc.getvalue(), indentation='   ', indent_text=True)

with open('Tanya.xml', 'w', encoding='UTF-8') as f:
    f.write(result)
